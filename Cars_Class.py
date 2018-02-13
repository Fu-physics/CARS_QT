
# Setup our imports
import sys
import time
import numpy as np
import scipy
import scipy.signal
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl import load_workbook
import als_methods as als
import pre as crikit


class CARS():
    def __init__(self):
        pass

    def nm2wavenumber(self, pump, antistokes):
        return 10**7*(1/np.array(antistokes) - 1/pump)
        
    def area(self, x):
        return np.sum(x)
    
    def is_number(self, s):
        try:
            float(s)
            return True
        except ValueError:
            pass
    
        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass

        return False


    def loadData(self, fileName):
        ##  --------------------------------------               Import the xlsx file
        W = load_workbook(fileName)
        p = W.get_sheet_by_name(name = 'Sheet1')

        a=[]
        for row in p.iter_rows():
            for k in row:
                a.append(k.internal_value)

        dataNumer = 0
        for i in a:
            if  (not self.is_number(i)):
                dataNumer  = dataNumer + 1
            elif self.is_number(i):
                break
        print("dataNumer is: ", dataNumer)

        data = np.resize(a, [int(len(a)/dataNumer), dataNumer])
        del a
        dataNameList = data[0,:]
        print("The dictionary name list is:\n", dataNameList)

        ##   -------------------------           reshape the data to dictionary, and normalized the data which makes the area of each plot to unit 1;
        dic ={}
        for i in range(dataNumer):
            dic[dataNameList[i]] = data[1:,i].astype(np.float)

        dic_norm = {}  ## nornalized data dictionary, not included the 0 index which is the wavelength
        for i in dataNameList[1:]:
            #print(i)
            dic_norm[i] = dic[i]/self.area(dic[i])

        dic_norm["wavenumber"] = self.nm2wavenumber(1064, dic[dataNameList[0]])
        dic_norm[dataNameList[0]] = dic[dataNameList[0]]

        dic["wavenumber"] = dic_norm["wavenumber"]

        return  dic, dic_norm

    def plotPredata(self, Predata, preFigure):

        fig = preFigure

        WN = Predata["wavenumber"]
        averageIntensity =  0.0 * Predata["wavenumber"]
        
        plt.subplot(111)     ## plot background 
        
        numberOfData = 0; 
        for key, value in Predata.items():
            if (key != "wavenumber" and key != "wavelength") :
                print(key)
                print(Predata[key])
                averageIntensity += value
                numberOfData += 1
                plt.plot(WN, value, label = key)
                plt.legend()
                
        ## The average of background_list, which will be used as the background data.
        averageIntensity = averageIntensity / numberOfData     
        
        plt.plot(WN, averageIntensity, label = 'averageIntensity')

        plt.legend()
        plt.xlabel('Wavenumber (cm$^{-1}$)')
        plt.ylabel('Spectrum (au)')
        plt.show()
    
class KK_ALS_Spectral():

    def __init__(self, signalDivBackground):
        self.I_CARS = signalDivBackground
        self.I_REF = np.zeros(len(signalDivBackground)) + 1
        self.Kramers_Kronig()

    def Kramers_Kronig(self):

        PHASE_OFFSET = 0 # DC phase-offset (default = 0)
        NORM_BY_NRB = 1 # Normalize retrieved spectrum by NRB/REF-- Removes the optical system response (default = 1)

        # Perform the Kramers-Kronig Relationship when using a surrogate NRB "reference"
        self.Retrieved_complex_spectrum_w_reference = crikit.kkrelation(self.I_REF,self.I_CARS, PHASE_OFFSET,NORM_BY_NRB) # Complex spectrum
        self.Retrieved_Raman_spectrum_w_reference = self.Retrieved_complex_spectrum_w_reference.imag # "Raman-like# (imag{complex spectrum})

        self.Retrieved_complex_spectrum_w_reference.imag = - self.Retrieved_complex_spectrum_w_reference.imag;
    

    def ALS(self):
        
        SMOOTHNESS_PARAM = 1e3
        ASYM_PARAM = 1e-4
        [Baseline, als_method] = als.als_baseline(np.angle(self.Retrieved_complex_spectrum_w_reference), SMOOTHNESS_PARAM, \
                                             ASYM_PARAM)

        # Error phase (baseline) and amplitude connected via Kramers-Kronig relation (Hilbert transform)
        Error_phase = Baseline
        Error_amp = np.exp(crikit.hilbertfft(Error_phase).imag)
        Correction_factor_1 = 1/Error_amp * np.exp(-1j*Error_phase)

        Phase_Corrected = self.Retrieved_complex_spectrum_w_reference*Correction_factor_1

        # Amiguity in phase-amp correction found by looking at mean-trend line of real componenet
        # Using low-order, large-window Savitky-Golay
        SGOLAY_WINDOW = 601
        SGOLAY_POLY_ORDER = 2

        Scaling_factor = 1/(scipy.signal.savgol_filter(np.real(Phase_Corrected),SGOLAY_WINDOW,SGOLAY_POLY_ORDER,axis=0))

        Corrected = Scaling_factor*Phase_Corrected

        # Compare these results with traditional baseline detrending of the Raman-like spectrum retrieved
        # directly from the Kramers-Kronig relation
        Just_amplitude_corrected = self.Retrieved_Raman_spectrum_w_reference - als.als_baseline(self.Retrieved_Raman_spectrum_w_reference,1e3,1e-4)[0]


        """
        These are defined earlier:
        SMOOTHNESS_PARAM
        ASYM_PARAM
        SGOLAY_WINDOW
        SGOLAY_POLY_ORDER
        """
        # Starts here
        Error_phase = als.als_baseline(np.angle(self.Retrieved_complex_spectrum_w_reference), SMOOTHNESS_PARAM,ASYM_PARAM)[0]
        Phase_Corrected = self.Retrieved_complex_spectrum_w_reference*(1/np.exp(crikit.hilbertfft(Error_phase).imag) * \
                                                                np.exp(-1j*Error_phase))
        Corrected = 1/(scipy.signal.savgol_filter(np.real(Phase_Corrected),601,2,axis=0))*Phase_Corrected
        # Done

        Just_amplitude_corrected = self.Retrieved_Raman_spectrum_w_reference - \
            als.als_baseline(self.Retrieved_Raman_spectrum_w_reference,1e3,1e-4)[0]

        return Corrected.imag, Just_amplitude_corrected




if __name__ == '__main__':
    app = CARS()
    dicBackground, dicBackgroundNorm = app.loadData('spectra_background.xlsx')
    dicSignal, dicSignalNorm = app.loadData('spectra_signal.xlsx')


    #fig = plt.figure(figsize=[12, 6])
    #app.plotPredata(dicSignalNorm, fig)

    signalDivBackground = dicBackgroundNorm['background_1']/dicSignalNorm['signal_1']
    print("signalDivBackground is: ", signalDivBackground)

    specturmMethod = KK_ALS_Spectral(signalDivBackground)
    spect_1, spect_2 = specturmMethod.ALS()

    plt.figure(figsize=[10, 4])
    plt.rc('font',size=12)
    plt.plot(spect_1, label='Phase Corrected + Scaling')
    plt.plot(spect_2, label = 'Baseline Detrended (only)')
    #plt.plot(WN,CHI_R.imag/np.abs(CHI_NR), 'r--', label = 'Ideal')
    plt.xlabel('Wavenumber (cm$^{-1}$)')
    plt.ylabel('Intensity (au)')
    plt.title('Spectrum', fontsize = 14)
    plt.legend(frameon = False)
    plt.show()

    #app.plotPredata(dicBackgroundNorm, fig)
    #print("dic is:", dic)
    #print("dic_normal is:", dic_norm)
    print("the last step!")