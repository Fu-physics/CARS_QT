
# coding: utf-8

# <img align=left src="CRIkit_Logo.png"> 
# <br>
# <br>
# <br>
# <br>
# <br>
# #####Coherent Raman Imaging toolKit

# Setup our imports
import sys
import time
import numpy as np
import scipy
import scipy.signal
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl import load_workbook
import als_methods as als
import pre as crikit


##  --------------------------------------               Import the xlsx file
W = load_workbook('spectra_data.xlsx')
p = W.get_sheet_by_name(name = 'Sheet1')

a=[]

for row in p.iter_rows():
    for k in row:
        a.append(k.internal_value)

data = np.resize(a, [int(len(a)/15), 15])
print("The dictionary name list is:\n", data[0,:])


##   --------------------------                       Function defination
def nm2wavenumber(pump, antistokes):
    return 10**7*(1/np.array(antistokes) - 1/pump)
def area(x):
    return np.sum(x)


##   -------------------------                       reshape the data to dictionary, and normalized the data which makes the area of each plot to unit 1;
dic ={}
for i in range(15):
    dic[data[0,i]] = data[1:2001,i].astype(np.float)


dic_norm = {}  ## nornalized data dictionary
for i in dic:
    #print(area(dic[i]))
    dic_norm[i] = dic[i]/area(dic[i])
dic_norm["wavenumber"] = nm2wavenumber(1064, dic[data[0,0]])

WN = dic_norm["wavenumber"]   ## wavenumber


              ## the data list that we are intrested in, after delete bad data.
background_list = ["background_1", "background_2", "background_3", "background_4", "background_5", 
 "background_6", "background_7", "background_8", "background_9"]

signal_list = ["signal_1", "signal_2", "signal_3", "signal_4", "signal_5"]

background =  0.0*dic_norm["background_1"]
for i in background_list:
    background += dic_norm[i]
background = background/len(background_list)     ## The average of background_list, which will be used as the background data.

signal =  0.0*dic_norm["signal_1"]
for i in signal_list:
    signal += dic_norm[i]
signal = signal/len(signal_list)      ## The average of signal_list, which will be used as the signal data.

print("Number of background data is:", len(background_list))
print("background is:", background)
print("Number of signal data is:", len(signal_list))
print("signal is:", signal)

##   ---------------------------------------------  plot background and signal 
fig = plt.figure(figsize=[12, 6])

plt.subplot(221)     ## plot background 
for i in background_list:
    plt.plot(WN, dic_norm[i], label = i)
    
plt.plot(WN, background, label = 'background')

plt.legend()
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Spectrum (au)')
plt.title('Background_list', fontsize=14)


plt.subplot(222)      ##plot signal
for i in signal_list:
    plt.plot(WN, dic_norm[i], label = i)

plt.plot(WN, signal, label = 'signal')

plt.legend()
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Spectrum (au)')
plt.title('Signal_list', fontsize=14)

plt.subplot(223)      ##plot Average signal and Average background
plt.plot(WN, signal, label = 'Average signal')
plt.plot(WN, background, label = 'Average background')

plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Spectrum (au)')
plt.legend()


plt.subplot(224)      ##plot signal/background
plt.plot(WN, signal/background, label = 'signal/background')
plt.plot(WN, 2000*(signal - background)+1, label = '2000*(signal - background)+1')

plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Spectrum (au)')
plt.legend()

fig.savefig('Signal and Background.png')

##--------------------------------------------------------------------------------------------------------
##----------------------------------    CARS signal procession    ----------------------------------------

# ###Simulate a CARS Signal

I_CARS = signal/background
I_REF = np.zeros(len(signal)) +1


PHASE_OFFSET = 0 # DC phase-offset (default = 0)
NORM_BY_NRB = 1 # Normalize retrieved spectrum by NRB/REF-- Removes the optical system response (default = 1)

# Perform the Kramers-Kronig Relationship when using a surrogate NRB "reference"
Retrieved_complex_spectrum_w_reference = crikit.kkrelation(I_REF,I_CARS,PHASE_OFFSET,NORM_BY_NRB) # Complex spectrum
Retrieved_Raman_spectrum_w_reference = Retrieved_complex_spectrum_w_reference.imag # "Raman-like# (imag{complex spectrum})

Retrieved_complex_spectrum_w_reference.imag = -Retrieved_complex_spectrum_w_reference.imag


# ### Phase-Detrending Using Asymmetric Least Square (ALS)
# ALS takes two parameters-- their ratio is the most important
SMOOTHNESS_PARAM = 1e3
ASYM_PARAM = 1e-4

fig = plt.figure(figsize=[12, 6])

plt.subplot(221)
plt.rc('font',size=12)
plt.plot(WN, I_CARS, label = 'CARS')
plt.plot(WN, I_REF, label = 'Ref NRB')

plt.legend(loc=0,frameon = False)
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Spectral Intensity (au)')

plt.subplot(223)
plt.plot(WN, np.angle(Retrieved_complex_spectrum_w_reference), label = 'w/NRB Ref')
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Phase (rad)')
plt.title('Retrieved Phase', fontsize = 14)
plt.legend(loc = 0, frameon = False)

# Find baseline with optimal als method (see CRIkit.utils.als_methods for more info)
[Baseline, als_method] = als.als_baseline(np.angle(Retrieved_complex_spectrum_w_reference), SMOOTHNESS_PARAM,                                              ASYM_PARAM)
plt.subplot(224)
plt.plot(WN, np.angle(Retrieved_complex_spectrum_w_reference), label = 'w/NRB Ref')
plt.plot(WN, Baseline, label = 'ALS Baseline')

plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Phase (rad)')
plt.title('Retrieved Phase and Baseline (' + als_method + ' Method)', fontsize = 14)
plt.legend(frameon = False)

fig.savefig('Phase-Detrending Using Asymmetric Least Square (ALS).png')



# ###Correcting for Phase and Amplitude Error
# Error phase (baseline) and amplitude connected via Kramers-Kronig relation (Hilbert transform)
Error_phase = Baseline
Error_amp = np.exp(crikit.hilbertfft(Error_phase).imag)
Correction_factor_1 = 1/Error_amp * np.exp(-1j*Error_phase)

Phase_Corrected = Retrieved_complex_spectrum_w_reference*Correction_factor_1

# Amiguity in phase-amp correction found by looking at mean-trend line of real componenet
# Using low-order, large-window Savitky-Golay
SGOLAY_WINDOW = 601
SGOLAY_POLY_ORDER = 2

Scaling_factor = 1/(scipy.signal.savgol_filter(np.real(Phase_Corrected),SGOLAY_WINDOW,SGOLAY_POLY_ORDER,axis=0))

Corrected = Scaling_factor*Phase_Corrected

# Compare these results with traditional baseline detrending of the Raman-like spectrum retrieved
# directly from the Kramers-Kronig relation
Just_amplitude_corrected = Retrieved_Raman_spectrum_w_reference - als.als_baseline(Retrieved_Raman_spectrum_w_reference,1e3,1e-4)[0]

fig = plt.figure(figsize=[10, 4])
plt.rc('font',size=12)
plt.plot(WN,Corrected.imag, label='Phase Corrected + Scaling')
plt.plot(WN,Just_amplitude_corrected, label = 'Baseline Detrended (only)')
#plt.plot(WN,CHI_R.imag/np.abs(CHI_NR), 'r--', label = 'Ideal')
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Intensity (au)')
plt.title('Corrected Raman-Like Spectrum', fontsize = 14)
plt.legend(frameon = False)

fig.savefig('Corrected Raman-Like Spectrum_1.png')





"""
These are defined earlier:
SMOOTHNESS_PARAM
ASYM_PARAM
SGOLAY_WINDOW
SGOLAY_POLY_ORDER
"""
# Starts here
Error_phase = als.als_baseline(np.angle(Retrieved_complex_spectrum_w_reference), SMOOTHNESS_PARAM,ASYM_PARAM)[0]
Phase_Corrected = Retrieved_complex_spectrum_w_reference*(1/np.exp(crikit.hilbertfft(Error_phase).imag) *                                                           np.exp(-1j*Error_phase))
Corrected = 1/(scipy.signal.savgol_filter(np.real(Phase_Corrected),601,2,axis=0))*Phase_Corrected
# Done

Just_amplitude_corrected = Retrieved_Raman_spectrum_w_reference - als.als_baseline(Retrieved_Raman_spectrum_w_reference,1e3,1e-4)[0]

fig = plt.figure(figsize=[12, 6])
plt.subplot(211)
plt.rc('font',size=12)
plt.plot(WN,Corrected.imag, label='Phase Corrected + Scaling')
plt.plot(WN,Just_amplitude_corrected, label = 'Baseline Detrended (only)')
plt.ylabel('Intensity (au)')
plt.title('Corrected Raman-Like Spectrum', fontsize = 14)
plt.legend(frameon = False)

plt.subplot(212)      ##plot Average signal and Average background
plt.plot(WN, signal, label = 'Average signal')
plt.plot(WN, background, label = 'Average background')
plt.xlabel('Wavenumber (cm$^{-1}$)')
plt.ylabel('Intensity (au)')
plt.title('Average of signal and background Spectrum', fontsize = 10)
plt.legend(frameon = False)

fig.savefig('Corrected Raman-Like Spectrum.png')

#plt.show()




